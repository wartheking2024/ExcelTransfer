import pandas as pd
from openpyxl import load_workbook
import shutil
import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk

class StudentCardApp:
    def __init__(self, root):
        self.root = root
        self.root.title("学籍卡生成器")

        self.source_path = tk.StringVar()
        self.template_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.num_students = tk.IntVar(value=10)

        self.default_mapping = {
            "C3": "姓  名",
            "F3": "性别",
            "H3": "学籍号",
            "C4": "出生日期",
            "F4": "民族",
            "H4": "学号",
            "C5": "入学时间",
            "H5": "籍贯",
            "C6": "家庭地址",
            "B7": "父亲",
            "E7": "父亲电话",
            "I7": "父亲工作单位",
            "B8": "母亲",
            "E8": "母亲电话",
            "I8": "母亲工作单位",
            "D9": "",
            "C33": "评语1"
        }

        self.field_mapping = self.default_mapping.copy()
        self.editing_entry = None  # 当前正在编辑的 Entry
        self.pending_focus_handler = None  # 保存未提交的回调函数

        self.create_widgets()

    def create_widgets(self):
        frm = tk.Frame(self.root)
        frm.pack(padx=10, pady=10)

        def browse(var, file=True):
            path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")]) if file else filedialog.askdirectory()
            if path:
                var.set(path)
                if var == self.source_path:
                    self.load_mapping_from_file()

        tk.Label(frm, text="班级信息文件").grid(row=0, column=0)
        tk.Entry(frm, textvariable=self.source_path, width=40).grid(row=0, column=1)
        tk.Button(frm, text="选择", command=lambda: browse(self.source_path)).grid(row=0, column=2)

        tk.Label(frm, text="学籍卡模板").grid(row=1, column=0)
        tk.Entry(frm, textvariable=self.template_path, width=40).grid(row=1, column=1)
        tk.Button(frm, text="选择", command=lambda: browse(self.template_path)).grid(row=1, column=2)

        tk.Label(frm, text="输出文件夹").grid(row=2, column=0)
        tk.Entry(frm, textvariable=self.output_path, width=40).grid(row=2, column=1)
        tk.Button(frm, text="选择", command=lambda: browse(self.output_path, file=False)).grid(row=2, column=2)

        tk.Label(frm, text="学生人数").grid(row=3, column=0)
        tk.Entry(frm, textvariable=self.num_students, width=10).grid(row=3, column=1, sticky='w')

        tk.Button(frm, text="字段映射设置", command=self.edit_mapping).grid(row=4, column=1, pady=5)
        tk.Button(frm, text="开始生成", command=self.generate_cards).grid(row=5, column=1, pady=10)

    def get_mapping_file_path(self):
        if not self.source_path.get():
            return None
        base = os.path.splitext(os.path.basename(self.source_path.get()))[0]
        return os.path.join(os.path.dirname(self.source_path.get()), f"{base}.mapping.json")

    def load_mapping_from_file(self):
        mapping_file = self.get_mapping_file_path()
        if mapping_file and os.path.exists(mapping_file):
            try:
                with open(mapping_file, "r", encoding="utf-8") as f:
                    self.field_mapping = json.load(f)
            except Exception as e:
                messagebox.showwarning("警告", f"读取字段映射失败：{e}")
        else:
            self.field_mapping = self.default_mapping.copy()

    def edit_mapping(self):
        self.load_mapping_from_file()

        mapping_window = tk.Toplevel(self.root)
        mapping_window.title("字段映射设置")
        mapping_window.geometry("400x400")

        tree = ttk.Treeview(mapping_window, columns=("cell", "field"), show="headings", selectmode="browse")
        tree.heading("cell", text="单元格位置")
        tree.heading("field", text="字段名")
        tree.column("cell", width=100, anchor="center")
        tree.column("field", width=200, anchor="w")
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        for cell, field in self.field_mapping.items():
            tree.insert("", "end", values=(cell, field))

        def on_double_click(event):
            region = tree.identify("region", event.x, event.y)
            if region != "cell":
                return
            row_id = tree.identify_row(event.y)
            column = tree.identify_column(event.x)
            if not row_id or not column:
                return

            x, y, width, height = tree.bbox(row_id, column)
            value0 = tree.set(row_id, column[0:])
            entry = tk.Entry(mapping_window)
            entry.place(x=x + tree.winfo_x(), y=y + tree.winfo_y(), width=width, height=height)
            entry.insert(0, value0)
            entry.focus_set()
            self.editing_entry = entry

            def on_focus_out(event):
                new_value = entry.get().strip()
                if column == "#1":
                    all_cells = [tree.set(i, "cell") for i in tree.get_children() if i != row_id]
                    if new_value in all_cells:
                        messagebox.showwarning("重复", f"单元格位置 {new_value} 已存在！")
                        entry.destroy()
                        self.editing_entry = None
                        return
                    import re
                    if not re.match(r"^[A-Z]+\d+$", new_value):
                        messagebox.showwarning("格式错误", "单元格格式应如 C3")
                        entry.destroy()
                        self.editing_entry = None
                        return
                tree.set(row_id, column[0:], new_value)
                entry.destroy()
                self.editing_entry = None

            self.pending_focus_handler = on_focus_out
            entry.bind("<FocusOut>", on_focus_out)
            entry.bind("<Return>", lambda e: on_focus_out(None))

        tree.bind("<Double-1>", on_double_click)

        def add_row():
            cell = simpledialog.askstring("单元格", "请输入单元格位置，如 C3", parent=mapping_window)
            if not cell:
                return
            import re
            if not re.match(r"^[A-Z]+\d+$", cell):
                messagebox.showwarning("格式错误", "单元格格式应如 C3", parent=mapping_window)
                return
            existing_cells = [tree.set(i, "cell") for i in tree.get_children()]
            if cell in existing_cells:
                messagebox.showwarning("重复", f"单元格位置 {cell} 已存在！", parent=mapping_window)
                return
            field = simpledialog.askstring("字段", "请输入字段名，如 姓  名", parent=mapping_window)
            if field is None:
                field = ""
            tree.insert("", "end", values=(cell, field))

        def delete_row():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请先选择一行", parent=mapping_window)
                return
            for item in selected:
                tree.delete(item)

        def save_and_close():
            if self.editing_entry and self.pending_focus_handler:
                self.pending_focus_handler(None)
                self.editing_entry = None

            new_mapping = {}
            for i in tree.get_children():
                cell = tree.set(i, "cell").strip()
                field = tree.set(i, "field").strip()
                if cell:
                    new_mapping[cell] = field
            self.field_mapping = new_mapping

            mapping_file = self.get_mapping_file_path()
            if mapping_file:
                try:
                    with open(mapping_file, "w", encoding="utf-8") as f:
                        json.dump(self.field_mapping, f, ensure_ascii=False, indent=2)
                except Exception as e:
                    messagebox.showerror("保存失败", str(e))

            mapping_window.destroy()

        def cancel_and_close():
            mapping_window.destroy()

        btn_frame = tk.Frame(mapping_window)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Button(btn_frame, text="添加", command=add_row).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="删除", command=delete_row).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="不保存退出", command=cancel_and_close).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_frame, text="保存并关闭", command=save_and_close).pack(side=tk.RIGHT, padx=5)

    def generate_cards(self):
        try:
            self.load_mapping_from_file()  # 确保字段映射是最新的

            df = pd.read_excel(self.source_path.get(), sheet_name='Sheet1')
            df = df.head(self.num_students.get())
            os.makedirs(self.output_path.get(), exist_ok=True)

            for idx, row in df.iterrows():
                student_name = row.get("姓  名", f"学生{idx+1}")
                output_file = os.path.join(self.output_path.get(), f"{student_name}_学籍卡.xlsx")
                shutil.copy(self.template_path.get(), output_file)

                wb = load_workbook(output_file)
                ws = wb.active

                for cell, column in self.field_mapping.items():
                    #print(cell)
                    #print(column)
                    if column == "":
                        ws[cell] = ""
                    elif column == "入学时间" and pd.notna(row.get(column)):
                        date = pd.to_datetime(row[column])
                        ws[cell] = date.strftime("%Y-%m-%d")
                    elif column in row and pd.notna(row[column]):
                        ws[cell] = row[column]
                    else:
                        ws[cell] = ""

                wb.save(output_file)

            messagebox.showinfo("完成", f"成功生成 {self.num_students.get()} 份学籍卡。")
        except Exception as e:
            messagebox.showerror("错误", str(e))


if __name__ == '__main__':
    root = tk.Tk()
    app = StudentCardApp(root)
    root.mainloop()
