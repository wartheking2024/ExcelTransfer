import pandas as pd
from openpyxl import load_workbook
import sys
import shutil
import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from PIL import Image, ImageTk  # 需安装 pillow 库：pip install pillow

def resource_path(relative_path):
    try:
        # PyInstaller打包后会把资源放在临时文件夹
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


class StudentCardApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel自动生成器")

        self.source_path = tk.StringVar()
        self.template_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.num_students = tk.IntVar(value=10)

        self.field_mapping = {}
        self.editing_entry = None
        self.pending_focus_handler = None
        self.mapping_checked = False

        self.create_widgets()

    def create_widgets(self):
        frm = tk.Frame(self.root)
        frm.pack(padx=10, pady=10)

        def browse(var, file=True):
            path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")]) if file else filedialog.askdirectory()
            if path:
                var.set(path)

        tk.Label(frm, text="数据集Excel文件").grid(row=0, column=0)
        tk.Entry(frm, textvariable=self.source_path, width=40).grid(row=0, column=1)
        tk.Button(frm, text="选择", command=lambda: browse(self.source_path)).grid(row=0, column=2)

        tk.Label(frm, text="生成Excel模板").grid(row=1, column=0)
        tk.Entry(frm, textvariable=self.template_path, width=40).grid(row=1, column=1)
        tk.Button(frm, text="选择", command=lambda: browse(self.template_path)).grid(row=1, column=2)

        tk.Label(frm, text="输出文件夹").grid(row=2, column=0)
        tk.Entry(frm, textvariable=self.output_path, width=40).grid(row=2, column=1)
        tk.Button(frm, text="选择", command=lambda: browse(self.output_path, file=False)).grid(row=2, column=2)

        tk.Label(frm, text="生成数量(上到下)").grid(row=3, column=0)
        tk.Entry(frm, textvariable=self.num_students, width=10).grid(row=3, column=1, sticky='w')

        tk.Button(frm, text="自动识别", command=self.auto_detect_mapping).grid(row=4, column=0, pady=5)
        tk.Button(frm, text="字段映射设置", command=self.edit_mapping).grid(row=4, column=1, pady=5)

        self.generate_btn = tk.Button(frm, text="开始生成", command=self.generate_cards, state=tk.DISABLED)
        self.generate_btn.grid(row=5, column=1, pady=10)

        # “打赏作者”文本按钮，放同一行右侧一格
        label_donate = tk.Label(frm, text="联系作者", fg="blue", cursor="hand2")
        label_donate.grid(row=5, column=2, pady=10, padx=(10,0))
        label_donate.bind("<Button-1>", lambda e: self.show_qrcode())



    def show_qrcode(self):
        # 新窗口
        win = tk.Toplevel(self.root)
        win.title("打赏作者")
        win.geometry("300x450")

        try:
            # 加载图片
            from PIL import Image, ImageTk
            img_path = resource_path("Pic/20250702222529.jpg")
            img = Image.open(img_path)  # 这里改成你的二维码图片路径
            img = img.resize((250, 350))  # 调整大小适应窗口
            photo = ImageTk.PhotoImage(img)

            # 添加文字提示
            label_text = tk.Label(win, text="如果觉得好用，欢迎打赏支持开发 😊", font=("Arial", 12))
            label_text.pack(pady=(10, 5))

            label = tk.Label(win, image=photo)
            label.image = photo  # 防止被垃圾回收
            label.pack(padx=10, pady=10)

            # 添加文字提示
            label_text1 = tk.Label(win, text="作者QQ:1322075214", fg="blue", font=("Arial", 12))
            label_text1.pack(pady=(10, 5))

        except Exception as e:
            tk.Label(win, text=f"无法加载二维码图片:\n{e}").pack(padx=10, pady=10)

    def get_mapping_file_path(self):
        if not self.source_path.get():
            return None
        base_dir = os.path.dirname(self.source_path.get())
        base_name = os.path.splitext(os.path.basename(self.source_path.get()))[0]

        save_dir = os.path.join(base_dir, "Save")
        os.makedirs(save_dir, exist_ok=True)

        return os.path.join(save_dir, f"{base_name}.mapping.json")

    def auto_detect_mapping(self):
        try:
            if not self.source_path.get() or not self.template_path.get():
                messagebox.showwarning("警告", "请先选择数据集Excel文件和生成模板文件。")
                return

            df = pd.read_excel(self.source_path.get(), sheet_name="Sheet1")
            headers = df.columns.tolist()

            wb = load_workbook(self.template_path.get())
            ws = wb.active

            mapping = {}

            # 记录合并单元格字典：起始坐标 -> 合并区域所有坐标
            merged_dict = {}
            for merge_range in ws.merged_cells.ranges:
                coords = list(ws[merge_range.coord])
                flat = [cell for row in coords for cell in row]
                start_cell = flat[0]
                end_cell = flat[-1]
                merged_dict[start_cell.coordinate] = (start_cell, end_cell)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    val = str(cell.value).strip() if cell.value else ""
                    if val in headers:
                        # 判断是否为合并单元格起点
                        coord = cell.coordinate
                        if coord in merged_dict:
                            _, end_cell = merged_dict[coord]
                            base_col = end_cell.column
                            base_row = end_cell.row
                        else:
                            base_col = cell.column
                            base_row = cell.row

                        # 尝试右边
                        right = ws.cell(row=base_row, column=base_col + 1)
                        down = ws.cell(row=base_row + 1, column=base_col)
                        if right.value is None or str(right.value).strip() == "":
                            mapping[right.coordinate] = val
                        elif down.value is None or str(down.value).strip() == "":
                            mapping[down.coordinate] = val

            if not mapping:
                messagebox.showwarning("识别失败", "未能在模板中找到与匹配的字段。")
                return

            self.field_mapping = mapping
            mapping_file = self.get_mapping_file_path()
            if mapping_file:
                with open(mapping_file, "w", encoding="utf-8") as f:
                    json.dump(self.field_mapping, f, ensure_ascii=False, indent=2)

            messagebox.showinfo("成功", "字段映射自动识别成功，请点击 字段映射设置 查看并确认。")

        except Exception as e:
            messagebox.showerror("错误", f"自动识别失败：{e}")

    def edit_mapping(self):
        mapping_file = self.get_mapping_file_path()
        if mapping_file and os.path.exists(mapping_file):
            with open(mapping_file, "r", encoding="utf-8") as f:
                self.field_mapping = json.load(f)

        mapping_window = tk.Toplevel(self.root)
        mapping_window.title("字段映射设置")
        mapping_window.geometry("400x400")

        tree = ttk.Treeview(mapping_window, columns=("cell", "field"), show="headings", selectmode="browse")
        tree.heading("cell", text="单元格位置")
        tree.heading("field", text="字段名")
        tree.column("cell", width=100, anchor="center")
        tree.column("field", width=200, anchor="w")
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        for cell, field in self.field_mapping.items():
            tree.insert("", "end", values=(cell, field))

        def on_double_click(event):
            region = tree.identify("region", event.x, event.y)
            if region != "cell":
                return
            row_id = tree.identify_row(event.y)
            column = tree.identify_column(event.x)
            if not row_id or not column:
                return

            x, y, width, height = tree.bbox(row_id, column)
            value0 = tree.set(row_id, column[0:])
            entry = tk.Entry(mapping_window)
            entry.place(x=x + tree.winfo_x(), y=y + tree.winfo_y(), width=width, height=height)
            entry.insert(0, value0)
            entry.focus_set()
            self.editing_entry = entry

            def on_focus_out(event):
                new_value = entry.get().strip()
                if column == "#1":
                    all_cells = [tree.set(i, "cell") for i in tree.get_children() if i != row_id]
                    if new_value in all_cells:
                        messagebox.showwarning("重复", f"单元格位置 {new_value} 已存在！")
                        entry.destroy()
                        self.editing_entry = None
                        return
                    import re
                    if not re.match(r"^[A-Z]+\d+$", new_value):
                        messagebox.showwarning("格式错误", "单元格格式应如 C3")
                        entry.destroy()
                        self.editing_entry = None
                        return
                tree.set(row_id, column[0:], new_value)
                entry.destroy()
                self.editing_entry = None

            self.pending_focus_handler = on_focus_out
            entry.bind("<FocusOut>", on_focus_out)
            entry.bind("<Return>", lambda e: on_focus_out(None))

        tree.bind("<Double-1>", on_double_click)

        def add_row():
            cell = simpledialog.askstring("单元格", "请输入单元格位置，如 C3", parent=mapping_window)
            if not cell:
                return
            import re
            if not re.match(r"^[A-Z]+\d+$", cell):
                messagebox.showwarning("格式错误", "单元格格式应如 C3", parent=mapping_window)
                return
            existing_cells = [tree.set(i, "cell") for i in tree.get_children()]
            if cell in existing_cells:
                messagebox.showwarning("重复", f"单元格位置 {cell} 已存在！", parent=mapping_window)
                return
            field = simpledialog.askstring("字段", "请输入字段名，如 姓  名", parent=mapping_window)
            if field is None:
                field = ""
            tree.insert("", "end", values=(cell, field))

        def delete_row():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请先选择一行", parent=mapping_window)
                return
            for item in selected:
                tree.delete(item)

        def save_and_close():
            if self.editing_entry and self.pending_focus_handler:
                self.pending_focus_handler(None)
                self.editing_entry = None

            new_mapping = {}
            for i in tree.get_children():
                cell = tree.set(i, "cell").strip()
                field = tree.set(i, "field").strip()
                if cell:
                    new_mapping[cell] = field
            self.field_mapping = new_mapping

            mapping_file = self.get_mapping_file_path()
            if mapping_file:
                try:
                    with open(mapping_file, "w", encoding="utf-8") as f:
                        json.dump(self.field_mapping, f, ensure_ascii=False, indent=2)
                except Exception as e:
                    messagebox.showerror("保存失败", str(e))

            self.mapping_checked = True
            self.generate_btn.config(state=tk.NORMAL)
            mapping_window.destroy()

        def cancel_and_close():
            mapping_window.destroy()

        btn_frame = tk.Frame(mapping_window)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Button(btn_frame, text="添加", command=add_row).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="删除", command=delete_row).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="不保存退出", command=cancel_and_close).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_frame, text="保存并关闭", command=save_and_close).pack(side=tk.RIGHT, padx=5)

    def generate_cards(self):
        if not self.mapping_checked:
            messagebox.showwarning("未确认字段映射", "请先完成字段映射设置确认后再生成Excel文件。")
            return
        try:
            df = pd.read_excel(self.source_path.get(), sheet_name='Sheet1')
            df = df.head(self.num_students.get())
            os.makedirs(self.output_path.get(), exist_ok=True)

            # 标准化列名：去除所有列名的前后空格和中间多余空格
            #df.columns = [str(col).strip().replace(" ", "") for col in df.columns]

            # 判断是否有“姓名”列
            has_name_column = "姓名" in df.columns
            first_column_name = df.columns[0]

            for idx, row in df.iterrows():
                if has_name_column:
                    student_name = str(row["姓名"]).strip()
                else:
                    student_name = str(row[first_column_name]).strip()

                if not student_name:
                    student_name = f"学生{idx+1}"

                output_file = os.path.join(self.output_path.get(), f"{student_name}.xlsx")
                shutil.copy(self.template_path.get(), output_file)

                wb = load_workbook(output_file)
                ws = wb.active

                for cell, column in self.field_mapping.items():
                    #print(cell)
                    #print(column)
                    if column == "":
                        ws[cell] = ""
                    elif column in row and pd.notna(row[column]):
                        val = row[column]
                        #print(val)
                        if isinstance(val, pd.Timestamp):
                            if val.hour == 0 and val.minute == 0 and val.second == 0:
                                ws[cell] = val.strftime("%Y-%m-%d")
                            else:
                                ws[cell] = val.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            ws[cell] = val
                    else:
                        ws[cell] = ""

                wb.save(output_file)

            messagebox.showinfo("完成", f"成功生成 {self.num_students.get()} 份Excel文件。")
        except Exception as e:
            if 'read-only' in str(e):
                messagebox.showerror("错误", "合并的单元格非第一格无法写入，请检查现有的单元格位置")
            else:
                messagebox.showerror("错误", str(e))


if __name__ == '__main__':
    root = tk.Tk()
    app = StudentCardApp(root)
    root.mainloop()
